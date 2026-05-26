from __future__ import annotations

import argparse
import json
import time
from pathlib import Path
from typing import Any

import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from freelance_starter.paths import ensure_parent, project_path

OPEN_METEO_FORECAST_URL = "https://api.open-meteo.com/v1/forecast"

CURRENT_COLUMNS = [
    "city",
    "time",
    "temperature",
    "temperature_unit",
    "humidity",
    "humidity_unit",
    "wind_speed",
    "wind_speed_unit",
]
DAILY_COLUMNS = [
    "city",
    "date",
    "temperature_max",
    "temperature_min",
    "precipitation_sum",
]
ERROR_COLUMNS = ["city", "latitude", "longitude", "attempts", "error"]


def load_cities(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        raise FileNotFoundError(f"City config not found: {path}")
    cities = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(cities, list):
        raise ValueError("City config must be a list.")
    for city in cities:
        for field in ("name", "latitude", "longitude"):
            if field not in city:
                raise ValueError(f"City entry is missing field: {field}")
    return cities


def fetch_weather(
    city: dict[str, Any],
    *,
    session: Any = requests,
    timeout: int = 20,
    retries: int = 2,
    backoff_seconds: float = 1.0,
) -> dict[str, Any]:
    last_error: Exception | None = None
    attempts = retries + 1

    for attempt in range(attempts):
        try:
            response = session.get(
                OPEN_METEO_FORECAST_URL,
                params={
                    "latitude": city["latitude"],
                    "longitude": city["longitude"],
                    "current": "temperature_2m,relative_humidity_2m,wind_speed_10m",
                    "daily": "temperature_2m_max,temperature_2m_min,precipitation_sum",
                    "forecast_days": 3,
                    "timezone": "auto",
                },
                headers={"User-Agent": "python-freelance-starter/1.0"},
                timeout=timeout,
            )
            response.raise_for_status()
            payload = response.json()
            payload["city_name"] = city["name"]
            return payload
        except requests.RequestException as exc:
            last_error = exc
            if attempt < retries and backoff_seconds > 0:
                time.sleep(backoff_seconds * (attempt + 1))

    message = f"Failed to fetch weather for {city['name']} after {attempts} attempt(s)"
    raise RuntimeError(message) from last_error


def normalize_current(payload: dict[str, Any]) -> dict[str, Any]:
    current = payload.get("current", {})
    units = payload.get("current_units", {})
    return {
        "city": payload["city_name"],
        "time": current.get("time"),
        "temperature": current.get("temperature_2m"),
        "temperature_unit": units.get("temperature_2m"),
        "humidity": current.get("relative_humidity_2m"),
        "humidity_unit": units.get("relative_humidity_2m"),
        "wind_speed": current.get("wind_speed_10m"),
        "wind_speed_unit": units.get("wind_speed_10m"),
    }


def normalize_daily(payload: dict[str, Any]) -> list[dict[str, Any]]:
    daily = payload.get("daily", {})
    rows: list[dict[str, Any]] = []
    for index, date in enumerate(daily.get("time", [])):
        rows.append(
            {
                "city": payload["city_name"],
                "date": date,
                "temperature_max": value_at(daily.get("temperature_2m_max"), index),
                "temperature_min": value_at(daily.get("temperature_2m_min"), index),
                "precipitation_sum": value_at(daily.get("precipitation_sum"), index),
            }
        )
    return rows


def value_at(values: list[Any] | None, index: int) -> Any:
    if not values or index >= len(values):
        return None
    return values[index]


def build_weather_tables(
    cities: list[dict[str, Any]],
    *,
    session: Any = requests,
    timeout: int = 20,
    retries: int = 2,
    backoff_seconds: float = 1.0,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    current_rows: list[dict[str, Any]] = []
    daily_rows: list[dict[str, Any]] = []
    error_rows: list[dict[str, Any]] = []

    for city in cities:
        try:
            payload = fetch_weather(
                city,
                session=session,
                timeout=timeout,
                retries=retries,
                backoff_seconds=backoff_seconds,
            )
        except Exception as exc:
            error_rows.append(
                {
                    "city": city.get("name"),
                    "latitude": city.get("latitude"),
                    "longitude": city.get("longitude"),
                    "attempts": retries + 1,
                    "error": str(exc),
                }
            )
            continue

        current_rows.append(normalize_current(payload))
        daily_rows.extend(normalize_daily(payload))

    current_weather = pd.DataFrame(current_rows, columns=CURRENT_COLUMNS)
    daily_forecast = pd.DataFrame(daily_rows, columns=DAILY_COLUMNS)
    api_errors = pd.DataFrame(error_rows, columns=ERROR_COLUMNS)
    return current_weather, daily_forecast, api_errors


def build_report_summary(
    city_count: int,
    current_weather: pd.DataFrame,
    daily_forecast: pd.DataFrame,
    api_errors: pd.DataFrame,
) -> pd.DataFrame:
    successful_cities = len(current_weather)
    failed_cities = len(api_errors)
    return pd.DataFrame(
        [
            {"metric": "Configured cities", "value": city_count},
            {"metric": "Successful cities", "value": successful_cities},
            {"metric": "Failed cities", "value": failed_cities},
            {"metric": "Current weather rows", "value": len(current_weather)},
            {"metric": "Forecast rows", "value": len(daily_forecast)},
        ]
    )


def write_weather_report(
    current_weather: pd.DataFrame,
    daily_forecast: pd.DataFrame,
    api_errors: pd.DataFrame,
    output_path: Path,
    *,
    city_count: int | None = None,
) -> None:
    ensure_parent(output_path)
    suffix = output_path.suffix.lower()
    if suffix == ".xlsx":
        summary = build_report_summary(
            city_count if city_count is not None else len(current_weather) + len(api_errors),
            current_weather,
            daily_forecast,
            api_errors,
        )
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            summary.to_excel(writer, sheet_name="report_summary", index=False)
            current_weather.to_excel(writer, sheet_name="current_weather", index=False)
            daily_forecast.to_excel(writer, sheet_name="daily_forecast", index=False)
            api_errors.to_excel(writer, sheet_name="api_errors", index=False)
        format_weather_workbook(output_path)
    elif suffix == ".csv":
        current_weather.to_csv(output_path, index=False, encoding="utf-8-sig")
    else:
        raise ValueError("Output file must end with .csv or .xlsx")


def format_weather_workbook(workbook_path: Path) -> None:
    workbook = load_workbook(workbook_path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for column_cells in worksheet.columns:
            column_letter = get_column_letter(column_cells[0].column)
            longest = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            worksheet.column_dimensions[column_letter].width = min(max(longest + 2, 12), 32)

        if worksheet.max_row >= 2 and worksheet.max_column >= 1:
            table_name = "".join(part.capitalize() for part in worksheet.title.split("_"))
            table_ref = f"A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"
            table = Table(displayName=f"{table_name}Table", ref=table_ref)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            worksheet.add_table(table)

    workbook.save(workbook_path)


def parse_args() -> argparse.Namespace:
    default_cities = project_path("projects", "03_api_report_tool", "input", "cities.json")
    default_output = project_path(
        "projects", "03_api_report_tool", "output", "weather_report.xlsx"
    )

    parser = argparse.ArgumentParser(
        description="Call Open-Meteo API and generate a weather report."
    )
    parser.add_argument("--cities", type=Path, default=default_cities)
    parser.add_argument("--output", type=Path, default=default_output)
    parser.add_argument("--timeout", type=int, default=20)
    parser.add_argument("--retries", type=int, default=2)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    cities = load_cities(args.cities)
    current_weather, daily_forecast, api_errors = build_weather_tables(
        cities,
        timeout=args.timeout,
        retries=args.retries,
    )
    write_weather_report(
        current_weather,
        daily_forecast,
        api_errors,
        args.output,
        city_count=len(cities),
    )
    print(
        f"Weather report created for {len(current_weather)}/{len(cities)} city/cities: "
        f"{args.output}"
    )
    if not api_errors.empty:
        print(f"Warning: {len(api_errors)} city/cities failed. See api_errors sheet.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
