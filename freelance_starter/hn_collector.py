from __future__ import annotations

import argparse
import time
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import urlparse

import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from freelance_starter.paths import ensure_parent, project_path

HACKER_NEWS_SEARCH_URL = "https://hn.algolia.com/api/v1/search"

STORY_COLUMNS = [
    "keyword",
    "title",
    "source_domain",
    "url",
    "author",
    "points",
    "comments",
    "created_at",
]
ERROR_COLUMNS = ["keyword", "attempts", "error"]


def normalize_story(keyword: str, hit: dict[str, Any]) -> dict[str, Any]:
    url = hit.get("url") or hit.get("story_url") or ""
    return {
        "keyword": keyword,
        "title": hit.get("title") or hit.get("story_title") or "",
        "source_domain": urlparse(url).netloc,
        "url": url,
        "author": hit.get("author") or "",
        "points": hit.get("points") or 0,
        "comments": hit.get("num_comments") or 0,
        "created_at": hit.get("created_at") or "",
    }


def fetch_keyword(
    keyword: str,
    *,
    limit: int = 20,
    session: Any = requests,
    timeout: int = 20,
    retries: int = 2,
    backoff_seconds: float = 1.0,
) -> list[dict[str, Any]]:
    keyword = keyword.strip()
    if not keyword:
        raise ValueError("keyword cannot be empty")
    if limit < 1 or limit > 100:
        raise ValueError("limit must be between 1 and 100")

    attempts = retries + 1
    last_error: Exception | None = None
    for attempt in range(attempts):
        try:
            response = session.get(
                HACKER_NEWS_SEARCH_URL,
                params={"query": keyword, "tags": "story", "hitsPerPage": limit},
                headers={"User-Agent": "python-freelance-starter/1.0"},
                timeout=timeout,
            )
            response.raise_for_status()
            payload = response.json()
            hits = payload.get("hits", [])
            return [normalize_story(keyword, hit) for hit in hits[:limit]]
        except requests.RequestException as exc:
            last_error = exc
            if attempt < retries and backoff_seconds > 0:
                time.sleep(backoff_seconds * (attempt + 1))

    message = f"Failed to fetch public results for '{keyword}' after {attempts} attempt(s)"
    raise RuntimeError(message) from last_error


def collect_keywords(
    keywords: Iterable[str],
    *,
    limit: int = 20,
    session: Any = requests,
    timeout: int = 20,
    retries: int = 2,
    backoff_seconds: float = 1.0,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    cleaned_keywords = [keyword.strip() for keyword in keywords if keyword.strip()]
    if not cleaned_keywords:
        raise ValueError("No keywords provided.")

    rows: list[dict[str, Any]] = []
    error_rows: list[dict[str, Any]] = []
    for keyword in cleaned_keywords:
        try:
            rows.extend(
                fetch_keyword(
                    keyword,
                    limit=limit,
                    session=session,
                    timeout=timeout,
                    retries=retries,
                    backoff_seconds=backoff_seconds,
                )
            )
        except RuntimeError as exc:
            error_rows.append(
                {"keyword": keyword, "attempts": retries + 1, "error": str(exc)}
            )

    stories = pd.DataFrame(rows, columns=STORY_COLUMNS)
    api_errors = pd.DataFrame(error_rows, columns=ERROR_COLUMNS)
    return stories, api_errors


def summarize_keywords(stories: pd.DataFrame, keywords: Iterable[str]) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for keyword in dict.fromkeys(keyword.strip() for keyword in keywords if keyword.strip()):
        matches = stories[stories["keyword"] == keyword] if not stories.empty else stories
        rows.append(
            {
                "keyword": keyword,
                "stories_collected": len(matches),
                "total_points": int(matches["points"].sum()) if not matches.empty else 0,
                "total_comments": int(matches["comments"].sum()) if not matches.empty else 0,
            }
        )
    return pd.DataFrame(rows)


def read_keywords(path: Path) -> list[str]:
    if not path.exists():
        raise FileNotFoundError(f"Keyword file not found: {path}")
    return [
        line.strip()
        for line in path.read_text(encoding="utf-8").splitlines()
        if line.strip() and not line.strip().startswith("#")
    ]


def write_report(
    keyword_summary: pd.DataFrame,
    stories: pd.DataFrame,
    api_errors: pd.DataFrame,
    output_path: Path,
) -> None:
    ensure_parent(output_path)
    suffix = output_path.suffix.lower()
    if suffix == ".xlsx":
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            keyword_summary.to_excel(writer, sheet_name="keyword_summary", index=False)
            stories.to_excel(writer, sheet_name="stories", index=False)
            api_errors.to_excel(writer, sheet_name="api_errors", index=False)
        format_report_workbook(output_path)
    elif suffix == ".csv":
        stories.to_csv(output_path, index=False, encoding="utf-8-sig")
    else:
        raise ValueError("Output file must end with .csv or .xlsx")


def format_report_workbook(workbook_path: Path) -> None:
    workbook = load_workbook(workbook_path)
    header_fill = PatternFill("solid", fgColor="145A56")
    header_font = Font(color="FFFFFF", bold=True)

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for column_cells in worksheet.columns:
            column_letter = get_column_letter(column_cells[0].column)
            longest = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in column_cells
            )
            worksheet.column_dimensions[column_letter].width = min(
                max(longest + 2, 12), 48
            )

        if worksheet.max_row >= 2:
            table_name = "".join(part.capitalize() for part in worksheet.title.split("_"))
            table_ref = f"A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"
            table = Table(displayName=f"{table_name}Table", ref=table_ref)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium4",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            worksheet.add_table(table)

    summary_sheet = workbook["keyword_summary"]
    if summary_sheet.max_row >= 2:
        chart = BarChart()
        chart.title = "Stories Collected by Keyword"
        chart.y_axis.title = "Stories"
        chart.x_axis.title = "Keyword"
        chart.legend = None
        data = Reference(summary_sheet, min_col=2, min_row=1, max_row=summary_sheet.max_row)
        categories = Reference(summary_sheet, min_col=1, min_row=2, max_row=summary_sheet.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.height = 8
        chart.width = 15
        summary_sheet.add_chart(chart, "F2")

    stories_sheet = workbook["stories"]
    story_widths = {
        "A": 18,
        "B": 56,
        "C": 28,
        "D": 68,
        "E": 20,
        "F": 12,
        "G": 12,
        "H": 24,
    }
    for column_letter, width in story_widths.items():
        stories_sheet.column_dimensions[column_letter].width = width
    for row in range(2, stories_sheet.max_row + 1):
        stories_sheet.row_dimensions[row].height = 34
        for cell in stories_sheet[row]:
            cell.alignment = Alignment(
                vertical="top", wrap_text=cell.column in (2, 4)
            )

    workbook.save(workbook_path)


def parse_args() -> argparse.Namespace:
    default_keywords = project_path("projects", "02_public_data_collector", "input", "keywords.txt")
    default_output = project_path(
        "projects", "02_public_data_collector", "output", "hn_stories_report.xlsx"
    )

    parser = argparse.ArgumentParser(
        description="Collect public Hacker News search results for keywords."
    )
    parser.add_argument("--keyword", action="append", default=[])
    parser.add_argument("--keywords-file", type=Path, default=default_keywords)
    parser.add_argument("--limit", type=int, default=10)
    parser.add_argument("--output", type=Path, default=default_output)
    parser.add_argument("--timeout", type=int, default=20)
    parser.add_argument("--retries", type=int, default=2)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    keywords = args.keyword or read_keywords(args.keywords_file)
    stories, api_errors = collect_keywords(
        keywords,
        limit=args.limit,
        timeout=args.timeout,
        retries=args.retries,
    )
    keyword_summary = summarize_keywords(stories, keywords)
    write_report(keyword_summary, stories, api_errors, args.output)
    print(
        f"Collected {len(stories)} rows for {len(keywords)} keyword(s), "
        f"failed={len(api_errors)}: {args.output}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
