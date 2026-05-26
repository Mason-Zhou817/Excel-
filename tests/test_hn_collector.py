from __future__ import annotations

from typing import Any

import pytest
import requests
from openpyxl import load_workbook

from freelance_starter.hn_collector import (
    collect_keywords,
    fetch_keyword,
    summarize_keywords,
    write_report,
)


class FakeResponse:
    def __init__(self, payload: dict[str, Any]) -> None:
        self.payload = payload

    def raise_for_status(self) -> None:
        return None

    def json(self) -> dict[str, Any]:
        return self.payload


class FakeSession:
    def __init__(self) -> None:
        self.calls: list[dict[str, Any]] = []

    def get(self, url: str, **kwargs: Any) -> FakeResponse:
        self.calls.append({"url": url, **kwargs})
        return FakeResponse(
            {
                "hits": [
                    {
                        "title": "Python automation idea",
                        "url": "https://example.com/python",
                        "author": "alice",
                        "points": 42,
                        "num_comments": 7,
                        "created_at": "2026-01-01T00:00:00Z",
                    }
                ]
            }
        )


class FlakySession(FakeSession):
    def __init__(self) -> None:
        super().__init__()
        self.attempts = 0

    def get(self, url: str, **kwargs: Any) -> FakeResponse:
        self.attempts += 1
        if self.attempts == 1:
            raise requests.Timeout("temporary timeout")
        return super().get(url, **kwargs)


class FailingSession:
    def get(self, url: str, **kwargs: Any) -> FakeResponse:
        raise requests.Timeout("network timeout")


def test_fetch_keyword_normalizes_hacker_news_hits() -> None:
    session = FakeSession()

    rows = fetch_keyword("python", session=session, limit=1)

    assert rows == [
        {
            "keyword": "python",
            "title": "Python automation idea",
            "source_domain": "example.com",
            "url": "https://example.com/python",
            "author": "alice",
            "points": 42,
            "comments": 7,
            "created_at": "2026-01-01T00:00:00Z",
        }
    ]
    assert session.calls[0]["params"]["query"] == "python"


def test_collect_keywords_requires_at_least_one_keyword() -> None:
    with pytest.raises(ValueError, match="No keywords"):
        collect_keywords(["  "], session=FakeSession())


def test_fetch_keyword_retries_transient_timeout() -> None:
    session = FlakySession()

    rows = fetch_keyword("python", session=session, retries=1, backoff_seconds=0)

    assert session.attempts == 2
    assert rows[0]["title"] == "Python automation idea"


def test_collect_keywords_records_failed_keyword() -> None:
    stories, api_errors = collect_keywords(
        ["python"],
        session=FailingSession(),
        retries=0,
        backoff_seconds=0,
    )

    assert stories.empty
    assert api_errors.loc[0, "keyword"] == "python"
    assert api_errors.loc[0, "attempts"] == 1


def test_write_report_creates_summary_stories_and_error_sheets(tmp_path: Any) -> None:
    stories, api_errors = collect_keywords(["python"], session=FakeSession())
    summary = summarize_keywords(stories, ["python"])
    output_path = tmp_path / "public_report.xlsx"

    write_report(summary, stories, api_errors, output_path)

    workbook = load_workbook(output_path, read_only=True)
    assert workbook.sheetnames == ["keyword_summary", "stories", "api_errors"]
    assert summary.loc[0, "stories_collected"] == 1
