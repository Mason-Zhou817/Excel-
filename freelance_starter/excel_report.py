from __future__ import annotations

import argparse
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Font

from freelance_starter.paths import ensure_parent, project_path

REQUIRED_COLUMNS = {
    "order_id",
    "date",
    "customer",
    "product",
    "category",
    "region",
    "salesperson",
    "quantity",
    "unit_price",
}

ISSUE_COLUMNS = ["source_file", "source_row", "order_id", "issue"]


def load_order_tables(input_dir: Path) -> pd.DataFrame:
    paths = sorted(
        path
        for path in input_dir.iterdir()
        if path.suffix.lower() in {".csv", ".xlsx", ".xls"}
    )
    if not paths:
        raise ValueError(f"No CSV or Excel files found in {input_dir}")

    tables: list[pd.DataFrame] = []
    for path in paths:
        if path.suffix.lower() == ".csv":
            table = pd.read_csv(path)
        else:
            table = pd.read_excel(path)
        table["source_file"] = path.name
        tables.append(table)

    return pd.concat(tables, ignore_index=True)


def clean_orders(raw_orders: pd.DataFrame) -> pd.DataFrame:
    cleaned_orders, _ = clean_orders_with_quality_report(raw_orders)
    return cleaned_orders


def clean_orders_with_quality_report(
    raw_orders: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    orders = raw_orders.copy()
    orders.columns = [
        str(column).strip().lower().replace(" ", "_") for column in orders.columns
    ]

    missing = REQUIRED_COLUMNS - set(orders.columns)
    if missing:
        missing_text = ", ".join(sorted(missing))
        raise ValueError(f"Missing required columns: {missing_text}")

    orders["_source_row"] = orders.index + 2
    issues: list[dict[str, object]] = []

    text_columns = ["order_id", "customer", "product", "category", "region", "salesperson"]
    for column in text_columns:
        orders[column] = orders[column].astype(str).str.strip()

    orders["date"] = pd.to_datetime(orders["date"], errors="coerce")
    orders["quantity"] = pd.to_numeric(orders["quantity"], errors="coerce")
    orders["unit_price"] = pd.to_numeric(orders["unit_price"], errors="coerce")

    issue_checks = [
        (orders["order_id"].eq(""), "Missing order_id"),
        (orders["date"].isna(), "Invalid date"),
        (orders["quantity"].isna(), "Invalid quantity"),
        (orders["unit_price"].isna(), "Invalid unit_price"),
        (orders["quantity"].notna() & (orders["quantity"] <= 0), "Quantity must be positive"),
        (orders["unit_price"].notna() & (orders["unit_price"] < 0), "Unit price cannot be negative"),
    ]
    for mask, message in issue_checks:
        issues.extend(issue_rows(orders.loc[mask], message))

    orders = orders.dropna(subset=["order_id", "date", "quantity", "unit_price"])
    orders = orders[orders["order_id"] != ""]
    orders = orders[orders["quantity"] > 0]
    orders = orders[orders["unit_price"] >= 0]

    duplicate_mask = orders.duplicated(subset=["order_id"], keep="last")
    issues.extend(issue_rows(orders.loc[duplicate_mask], "Duplicate order_id, kept latest row"))
    orders = orders.drop_duplicates(subset=["order_id"], keep="last")

    orders["revenue"] = (orders["quantity"] * orders["unit_price"]).round(2)
    orders["month"] = orders["date"].dt.to_period("M").astype(str)
    orders = orders.sort_values(["date", "order_id"]).reset_index(drop=True)

    ordered_columns = [
        "order_id",
        "date",
        "month",
        "customer",
        "product",
        "category",
        "region",
        "salesperson",
        "quantity",
        "unit_price",
        "revenue",
        "source_file",
    ]
    existing_columns = [column for column in ordered_columns if column in orders.columns]
    quality_report = pd.DataFrame(issues, columns=ISSUE_COLUMNS)
    return orders[existing_columns], quality_report


def issue_rows(rows: pd.DataFrame, issue: str) -> list[dict[str, object]]:
    result: list[dict[str, object]] = []
    for _, row in rows.iterrows():
        result.append(
            {
                "source_file": row.get("source_file", ""),
                "source_row": row.get("_source_row", ""),
                "order_id": row.get("order_id", ""),
                "issue": issue,
            }
        )
    return result


def summarize_by_category(cleaned_orders: pd.DataFrame) -> pd.DataFrame:
    summary = (
        cleaned_orders.groupby("category", as_index=False)
        .agg(quantity=("quantity", "sum"), revenue=("revenue", "sum"))
        .sort_values("revenue", ascending=False)
    )
    summary["revenue"] = summary["revenue"].round(2)
    return summary


def summarize_by_month(cleaned_orders: pd.DataFrame) -> pd.DataFrame:
    monthly = (
        cleaned_orders.groupby("month", as_index=False)
        .agg(orders=("order_id", "count"), quantity=("quantity", "sum"), revenue=("revenue", "sum"))
        .sort_values("month")
    )
    monthly["revenue"] = monthly["revenue"].round(2)
    return monthly


def summarize_by_customer(cleaned_orders: pd.DataFrame) -> pd.DataFrame:
    customer = (
        cleaned_orders.groupby("customer", as_index=False)
        .agg(orders=("order_id", "count"), revenue=("revenue", "sum"))
        .sort_values("revenue", ascending=False)
    )
    customer["revenue"] = customer["revenue"].round(2)
    return customer


def write_excel_report(
    cleaned_orders: pd.DataFrame,
    quality_report: pd.DataFrame,
    category_summary: pd.DataFrame,
    monthly_summary: pd.DataFrame,
    customer_summary: pd.DataFrame,
    output_path: Path,
) -> None:
    ensure_parent(output_path)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        cleaned_orders.to_excel(writer, sheet_name="cleaned_orders", index=False)
        quality_report.to_excel(writer, sheet_name="data_quality", index=False)
        category_summary.to_excel(writer, sheet_name="summary_by_category", index=False)
        monthly_summary.to_excel(writer, sheet_name="monthly_trend", index=False)
        customer_summary.to_excel(writer, sheet_name="top_customers", index=False)

    format_workbook(output_path)


def format_workbook(workbook_path: Path) -> None:
    workbook = load_workbook(workbook_path)

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        for cell in worksheet[1]:
            cell.font = Font(bold=True)

    category_sheet = workbook["summary_by_category"]
    if category_sheet.max_row >= 2:
        chart = BarChart()
        chart.title = "Revenue by Category"
        chart.y_axis.title = "Revenue"
        chart.x_axis.title = "Category"

        data = Reference(category_sheet, min_col=3, min_row=1, max_row=category_sheet.max_row)
        categories = Reference(category_sheet, min_col=1, min_row=2, max_row=category_sheet.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.height = 8
        chart.width = 14
        category_sheet.add_chart(chart, "E2")

    monthly_sheet = workbook["monthly_trend"]
    if monthly_sheet.max_row >= 2:
        line_chart = LineChart()
        line_chart.title = "Monthly Revenue Trend"
        line_chart.y_axis.title = "Revenue"
        line_chart.x_axis.title = "Month"
        data = Reference(monthly_sheet, min_col=4, min_row=1, max_row=monthly_sheet.max_row)
        categories = Reference(monthly_sheet, min_col=1, min_row=2, max_row=monthly_sheet.max_row)
        line_chart.add_data(data, titles_from_data=True)
        line_chart.set_categories(categories)
        line_chart.height = 8
        line_chart.width = 14
        monthly_sheet.add_chart(line_chart, "F2")

    workbook.save(workbook_path)


def build_report(input_dir: Path, output_path: Path, summary_csv_path: Path | None) -> dict[str, int]:
    raw_orders = load_order_tables(input_dir)
    cleaned_orders, quality_report = clean_orders_with_quality_report(raw_orders)
    category_summary = summarize_by_category(cleaned_orders)
    monthly_summary = summarize_by_month(cleaned_orders)
    customer_summary = summarize_by_customer(cleaned_orders)
    write_excel_report(
        cleaned_orders,
        quality_report,
        category_summary,
        monthly_summary,
        customer_summary,
        output_path,
    )

    if summary_csv_path is not None:
        ensure_parent(summary_csv_path)
        category_summary.to_csv(summary_csv_path, index=False, encoding="utf-8-sig")

    return {
        "source_rows": len(raw_orders),
        "clean_rows": len(cleaned_orders),
        "quality_issues": len(quality_report),
        "categories": len(category_summary),
    }


def parse_args() -> argparse.Namespace:
    default_input = project_path("projects", "01_excel_report_automation", "input")
    default_output = project_path(
        "projects", "01_excel_report_automation", "output", "sales_report.xlsx"
    )
    default_summary = project_path(
        "projects", "01_excel_report_automation", "output", "category_summary.csv"
    )

    parser = argparse.ArgumentParser(
        description="Merge sales files, clean data, and build an Excel report."
    )
    parser.add_argument("--input-dir", type=Path, default=default_input)
    parser.add_argument("--output", type=Path, default=default_output)
    parser.add_argument("--summary-csv", type=Path, default=default_summary)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    stats = build_report(args.input_dir, args.output, args.summary_csv)
    print(
        "Report created: "
        f"{args.output} | source rows={stats['source_rows']} "
        f"clean rows={stats['clean_rows']} "
        f"quality issues={stats['quality_issues']} "
        f"categories={stats['categories']}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
